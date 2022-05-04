from ckiptagger import WS, POS, NER
from openpyxl import Workbook, load_workbook
from pathlib import Path
import operator

gwd = Path.cwd()

def tw_datalist():
    path = gwd / "twdata" / "twdata.xlsx"
    twb = load_workbook(path)
    tws = twb.worksheets[0]
    temp_list = []
    for i in range(tws.max_row-2):
        data = tws.cell(row=i+3,column=6).value
        temp_list.append(data)
    twb.close()
    return temp_list

def handle_twformat(x,y):
    all = list(zip(x,y))
    filter_list = ["Na","Nb","Nc","Ncd","Nd","A","Nv","VA","VAC","VB","VC","VCL","VD","VE","VF","VG","VH","VHC"]
    new_word_list = []
    for i in all:
        if i[1] in filter_list:
            new_word_list.append(i)
    return new_word_list

def count_keyword(word_list):
    dic = {}
    for ele in word_list:
        if ele not in dic:
            dic[ele] = 1
        else:
            dic[ele] = dic[ele] + 1
    sorted_dic = sorted(dic.items(),key=operator.itemgetter(1),reverse=True)
    return sorted_dic

def write_twkeyword(sorted_dic):
    twb = Workbook()
    tws = twb.active
    path = gwd / "twdata" / "twkeywords.xlsx"
    base_row = 1
    for ele in sorted_dic:
        tws.cell(row=base_row,column=1).value = str(ele[0][0])
        tws.cell(row=base_row,column=2).value = str(ele[1])
        tws.cell(row=base_row,column=3).value = str(ele[0][1])
        base_row += 1
    twb.save(path)

twlist = tw_datalist() 
ws = WS("./data")
pos = POS("./data")
ner = NER("./data")
ws_results = ws(twlist,)
pos_results = pos(ws_results)
ws_results = [b for a in ws_results for b in a]
pos_results = [b for a in pos_results for b in a]
new_word_list = handle_twformat(ws_results,pos_results)
sorted_dic = count_keyword(new_word_list)
write_twkeyword(sorted_dic)