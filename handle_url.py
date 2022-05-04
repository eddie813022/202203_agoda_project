from openpyxl import load_workbook
from collections import OrderedDict

uwb = load_workbook("urls.xlsx")
uws = uwb.active

url_list = []
for i in range(uws.max_row):
    url = uws.cell(row=i+1,column=1).value
    index = url.find("?final")
    url = url[:index]
    url_list.append(url)

new_list = list(OrderedDict.fromkeys(url_list))
for i in range(len(new_list)):
    uws.cell(row=i+1,column=1).value = new_list[i]

uwb.save("urls.xlsx")    