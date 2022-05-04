from sklearn.feature_extraction.text import TfidfVectorizer
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd

# check headdata
# pd.set_option('display.max_rows', None)
# pd.set_option('display.max_columns', None)
# pd.set_option('display.width', None)
# pd.set_option('display.max_colwidth', -1)

gwd = Path.cwd()

def get_kr_priciple():
    path = gwd / "krdata" / "krdata.xlsx"
    wb = load_workbook(path)
    ws = wb.worksheets[1]
    dicts = {}
    for i in range(ws.max_row):
        key = ws.cell(row=i+1,column=1).value
        value = ws.cell(row=i+1,column=3).value
        dicts[key] = value
    return dicts

def get_tw_priciple():
    path = gwd / "twdata" / "twdata.xlsx"
    wb = load_workbook(path)
    ws = wb.worksheets[1]
    dicts = {}
    for i in range(ws.max_row):
        key = ws.cell(row=i+1,column=1).value
        value = ws.cell(row=i+1,column=3).value
        dicts[key] = value
    return dicts        

def handle_tw_predata():
    path_twdata = gwd / "twdata" / "twdata.xlsx"
    path_participle = gwd / "twdata" / "tw_participle.xlsx"
    twb = load_workbook(path_twdata)
    pwb = load_workbook(path_participle)
    pws = pwb.active
    twbs = twb.active
    basee = 3
    tw_temp_list = []
    tw_index_temp_list = []
    for i in range(twbs.max_row-2):
        value = twbs.cell(row=basee,column=6).value
        tw_index_temp_list.append(value)
        basee += 1
    for i in range(pws.max_row):
        value = pws.cell(row=i+1,column=1).value
        tw_temp_list.append(value)
    twb.close()
    pwb.close()
    return tw_temp_list,tw_index_temp_list

def handle_kr_predata():
    path = gwd / "krdata" / "krdata.xlsx"
    krb = load_workbook(path)
    krbs = krb.active
    base = 1
    kr_temp_list = []
    kr_index_temp_list = []
    for i in range(krbs.max_row):
        value = krbs.cell(row=base,column=9).value
        index_value = krbs.cell(row=base,column=7).value
        kr_index_temp_list.append(index_value)
        kr_temp_list.append(value)
        base += 1
    krb.close()
    return kr_temp_list,kr_index_temp_list

def handle_jp_predata():
    path = gwd / "jpdata.xlsx"
    jwb = load_workbook(path)
    jws = jwb.worksheets[0]
    jp_temp_list = []
    jp_index_temp_list = []
    for i in range(jws.max_row-2):
        value = jws.cell(row=i+3,column=9).value
        index_value = jws.cell(row=i+3,column=6).value
        jp_temp_list.append(value)
        jp_index_temp_list.append(index_value)
    jwb.close()
    return jp_temp_list,jp_index_temp_list
        
def handle_tf_idf(predata,index_data):
    vectoriser = TfidfVectorizer(norm = None)
    tf_idf_scores = vectoriser.fit_transform(predata)
    feature_names = vectoriser.get_feature_names_out()
    df_tf_idf = pd.DataFrame(tf_idf_scores.T.todense(), index = feature_names, columns = index_data)
    return df_tf_idf

def handle_sum(df):
    newdf = df.T
    sum_ = newdf.sum()
    return sum_

def handle_mean(df):
    newdf = df.T
    mean_ = newdf.mean()
    return mean_
    
def handle_max(df):
    newdf = df.T
    max_ = newdf.max()
    return max_

def handle_jp_sum():
    temp_list,index_list = handle_jp_predata()
    df = handle_tf_idf(temp_list,index_list)
    sum_df = handle_sum(df)
    sum_df.to_csv("sum.csv",encoding="utf-8_sig")

def handle_kr_sum():
    records = []
    df = pd.read_csv("krdata\kr_sum.csv")
    dicts = get_kr_priciple()
    names = df["word"]
    for (colname,colval) in names.iteritems():
        records.append(dicts[colval])
    df.insert(2,column="type",value=records)
    df.to_csv("sum.csv",encoding="utf-8_sig")

def handle_tw_sum():
    records = []
    df = pd.read_csv(r"twdata\tw_sum.csv")
    dicts = get_tw_priciple()
    for index,row in df.iterrows():
        string_ = dicts.get(row["word"])
        if string_:
            records.append(string_)
        else:
            df.drop(index=index,inplace=True)
    df.insert(2,column="type",value=records)
    df.reset_index(inplace=True)
    df.to_csv("sum.csv",encoding="utf-8_sig")





