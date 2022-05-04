from openpyxl import Workbook, load_workbook
from pathlib import Path
import operator
import MeCab

gwd = Path.cwd()

def jp_datalist():
    path = gwd / "jpdata" / "jpdata.xlsx"
    jwb = load_workbook(path)
    jws = jwb.worksheets[0]
    temp_list = []
    for i in range(jws.max_row-2):
        data = jws.cell(row=i+3,column=6).value
        temp_list.append(data)
    text = "".join(temp_list)
    jwb.close()
    return text

def count_keyword(word_list):
    dic = {}
    for ele in word_list:
        if ele not in dic:
            dic[ele] = 1
        else:
            dic[ele] = dic[ele] + 1
    sorted_dic = sorted(dic.items(),key=operator.itemgetter(1),reverse=True)
    return sorted_dic

def write_jpkeyword(sorted_dic):
    jwb = Workbook()
    jws = jwb.active
    path = gwd / "jpdata" / "jpkeywords.xlsx"
    base_row = 1
    for ele in sorted_dic:
        jws.cell(row=base_row,column=1).value = str(ele[0][0])
        jws.cell(row=base_row,column=2).value = str(ele[1])
        jws.cell(row=base_row,column=3).value = str(ele[0][1])
        base_row += 1
    jws.save(path)


m = MeCab.Tagger('-Osimple -d /etc/alternatives/mecab-dictionary')
statement = jp_datalist()
node = m.parseToNode(statement)
words=[]
while node:
    hinshi = node.feature.split(",")[0]
    if hinshi in ["名詞","動詞","形容詞"]:
        origin = node.feature.split(",")[6]
        newword = (origin,hinshi)
        words.append(newword)
    node = node.next
sorted_dic = count_keyword(words)
write_jpkeyword(sorted_dic)





