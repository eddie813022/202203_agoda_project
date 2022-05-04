from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import InvalidArgumentException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from decimal import Decimal,getcontext,DecimalException
from pathlib import Path
from time import sleep
import urllib3
import random

# ignore request warning 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# decimal configure
getcontext().prec

# path configure
cwd = Path.getcwd()

# workbook configure
uwb = load_workbook("urls.xlsx")
uws = uwb.active
tw_base_row = 3
jp_base_row = 3
kr_base_row = 3

# selenium configure
options = webdriver.FirefoxOptions()
options.add_argument("--headless")
options.add_argument("--disable-notifications")
options.add_argument('--ignore-certificate-errors')
drvier = cwd / "drivers" / "geckodriver.exe"
service = Service(drvier)
driver = webdriver.Firefox(service=service,options=options) 
driver.maximize_window()

# get_agoda_urlS
def get_agoda_urls():
    driver.get("https://www.agoda.com/zh-tw/search?city=14690&hotelStarRating=5,4&hotelAccom=34")
    driver.implicitly_wait(5)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    sleep(4)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    sleep(4)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    sleep(2)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    soup = BeautifulSoup(driver.page_source,"html.parser")
    hotel_url_list = soup.findAll("li",{"data-selenium":"hotel-item"})
    for hotels in hotel_url_list:
        for hotel_url in hotels.find_all("a",href=True):
            suffix_url = hotel_url["href"]
            full_url = "https://www.agoda.com"+suffix_url
            uws.append({1:full_url})
    driver.find_element(By.ID,"paginationNext").click()
    driver.implicitly_wait(3)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    soup2 = BeautifulSoup(driver.page_source,"html.parser")
    hotel_url_list2 = soup.findAll("li",{"data-selenium":"hotel-item"})
    for hotels in hotel_url_list:
        for hotel_url in hotels.find_all("a",href=True):
            suffix_url = hotel_url["href"]
            full_url = "https://www.agoda.com"+suffix_url
            uws.append({1:full_url})
    uwb.save("urls_1.xlsx")        

# sleep func
def ransleep():
    sleep(random.randint(2,3))

# write_comment func
def write_comments(soup,language):
    global tw_base_row,kr_base_row,jp_base_row
    hotel_name = soup.findAll("h1",class_="HeaderCerebrum__Name")[0].string
    if "(" in hotel_name:
        index = hotel_name.find("(")
        hotel_name = hotel_name[index+1:].replace(")","")
    hotel_score = soup.select_one("h3.Typographystyled__TypographyStyled-sc-j18mtu-0.hTkvyT.kite-js-Typography").text
    pcomment_soup = soup.findAll("div",class_="Review-comment")
    pscore_list = [ pcomment.findAll("div",class_="Review-comment-leftScore")[0].string for pcomment in pcomment_soup ]
    pcontent_list = [ pcomment.findAll("p",class_="Review-comment-bodyText")[0].string for pcomment in pcomment_soup ]
    pdate_pre_list = [ pcomment.findAll("span",class_="Review-statusBar-date")[0].string for pcomment in pcomment_soup ]
    pdate_list = [ "".join(pdate)[5:-3].replace("年","/").replace("月","/").replace("日","") for pdate in pdate_pre_list ]
    if pcontent_list:
        g = 0
        if language == "tw":
            for i in range(len(pcontent_list)):
                try:
                    ws_tw.cell(tw_base_row,1,url)
                    ws_tw.cell(tw_base_row,2,hotel_name)
                    ws_tw.cell(tw_base_row,3,hotel_score)
                    ws_tw.cell(tw_base_row,5,pscore_list[g])
                    ws_tw.cell(tw_base_row,6,len(pcontent_list[g]))
                    ws_tw.cell(tw_base_row,7,pcontent_list[g])
                    ws_tw.cell(tw_base_row,8,pdate_list[g])
                except:
                    pass
                g += 1
                tw_base_row += 1
        elif language == "kr":
            for i in range(len(pcontent_list)):
                try:
                    ws_kr.cell(kr_base_row,1,url)
                    ws_kr.cell(kr_base_row,2,hotel_name)
                    ws_kr.cell(kr_base_row,3,hotel_score)
                    ws_kr.cell(kr_base_row,5,pscore_list[g])
                    ws_kr.cell(kr_base_row,6,len(pcontent_list[g]))
                    ws_kr.cell(kr_base_row,7,pcontent_list[g])
                    ws_kr.cell(kr_base_row,8,pdate_list[g])
                except:
                    pass
                g += 1
                kr_base_row += 1
        elif language == "jp":
            for i in range(len(pcontent_list)):
                try:
                    ws_jp.cell(jp_base_row,1,url)
                    ws_jp.cell(jp_base_row,2,hotel_name)
                    ws_jp.cell(jp_base_row,3,hotel_score)
                    ws_jp.cell(jp_base_row,5,pscore_list[g])
                    ws_jp.cell(jp_base_row,6,len(pcontent_list[g]))
                    ws_jp.cell(jp_base_row,7,pcontent_list[g])
                    ws_jp.cell(jp_base_row,8,pdate_list[g])
                except:
                    pass
                g += 1
                kr_base_row += 1

# detect next page button
def count_page():
    try:
        ele_page_count = driver.find_element(By.CLASS_NAME,"Review__SummaryContainer")
    except:
        ele_page_count = ""
        page_list = 0
    if ele_page_count:
        ele_page_count = ele_page_count.find_element(By.CSS_SELECTOR,".Review__SummaryContainer--left.Review__SummaryContainer__Text").get_attribute("innerHTML")
        index_last = ele_page_count.find("篇")
        page_num = ele_page_count[1:index_last]
        if "," in page_num:
            page_num = page_num.replace(",","")
        try:
            count_page = (Decimal(page_num) / Decimal(20)).quantize(Decimal('.000')) 
            index = str(count_page).find(".")
            check_page = str(count_page)[index+1:]
            if check_page == "000":
                page_list =  int(count_page)
            else:
                page_list = int(count_page) + 1
        except DecimalException as e:
            print(e)
            page_list = 0
    else:
        page_list = 0
    return page_list

# page_click func
def page_click(language):
    for page in range((page_list-1)):
        ele = driver.find_element(By.CLASS_NAME,"Review-filters")
        driver.execute_script("arguments[0].scrollIntoView()",ele)
        next_page_btn = driver.find_element(By.CSS_SELECTOR,".Review-paginator[data-location='top']").find_element(By.CLASS_NAME,"Review-paginator-steps") \
        .find_elements(By.CSS_SELECTOR,"span")[-1]
        next_page_btn.click()
        sleep(1)
        WebDriverWait(driver,5).until(EC.presence_of_element_located((By.CSS_SELECTOR,".Review-paginator.Review-paginator--withSteps")))
        sleep(1)
        soup = BeautifulSoup(driver.page_source,"html.parser")
        write_comments(soup,language)
        ransleep()

# cralwer strings
for hotels in range(uws.max_row):
    wb = load_workbook("urls.xlsx")
    ws_tw = wb.worksheets[0]
    ws_jp = wb.worksheets[0]
    ws_kr = wb.worksheets[0]
    url = uws.cell(row=hotels+1,column=1).value
    try:
        driver.get(url)
    except InvalidArgumentException:
        print(url)
    WebDriverWait(driver,5).until(EC.presence_of_element_located((By.ID,"reviewSectionComments")))
    try:
        filter_btn = driver.find_element(By.CSS_SELECTOR,".Review__FilterContainer__Dropbox[data-selenium='reviews-language-filter']") \
        .find_element(By.CSS_SELECTOR,"span")
        languege_btn = driver.find_element(By.CSS_SELECTOR,".Review__FilterContainer__Dropbox[data-selenium='reviews-language-filter']") \
        .find_element(By.CSS_SELECTOR,".select-dropdown-list.select-dropdown-list--align-undefined.Review__FilterContainer__List.ReviewFilterList.Review__FilterContainer__List--withScroll")
    except:
        filter_btn = ""
        languege_btn = ""
        chinese_btn = ""
        korea_btn = ""
        japan_btn = ""
    if languege_btn:
        if "繁體中文 (台灣)" in languege_btn.get_attribute("innerHTML"):
            chinese_btn = driver.find_element(By.CSS_SELECTOR,".Review__FilterContainer__Dropbox[data-selenium='reviews-language-filter']") \
            .find_element(By.CSS_SELECTOR,".select-dropdown-list.select-dropdown-list--align-undefined.Review__FilterContainer__List.ReviewFilterList.Review__FilterContainer__List--withScroll") \
            .find_element(By.XPATH,".//span[text()='繁體中文 (台灣)']").find_element(By.XPATH,"..")
        else:
            chinese_btn = ""
        if "한국어" in languege_btn.get_attribute("innerHTML"):
            korea_btn = driver.find_element(By.CSS_SELECTOR,".Review__FilterContainer__Dropbox[data-selenium='reviews-language-filter']") \
            .find_element(By.CSS_SELECTOR,".select-dropdown-list.select-dropdown-list--align-undefined.Review__FilterContainer__List.ReviewFilterList.Review__FilterContainer__List--withScroll") \
            .find_element(By.XPATH,".//span[text()='한국어']").find_element(By.XPATH,"..")
        else:
            korea_btn = ""
        if "日本語" in languege_btn.get_attribute("innerHTML"):
            japan_btn = driver.find_element(By.CSS_SELECTOR,".Review__FilterContainer__Dropbox[data-selenium='reviews-language-filter']") \
            .find_element(By.CSS_SELECTOR,".select-dropdown-list.select-dropdown-list--align-undefined.Review__FilterContainer__List.ReviewFilterList.Review__FilterContainer__List--withScroll") \
            .find_element(By.XPATH,".//span[text()='日本語']").find_element(By.XPATH,"..")
        else:
            japan_btn = ""
    # scoll down page
    ele = driver.find_element(By.CLASS_NAME,"Review-filters")
    driver.execute_script("arguments[0].scrollIntoView()",ele)
    ransleep()
    # make sure windows normal
    try:
        ActionChains(driver).move_by_offset(200, 100).click().perform()
        sleep(1)
    except:
        pass
    
    if chinese_btn:
        filter_btn.click()
        chinese_btn.click()
        WebDriverWait(driver,5).until(EC.presence_of_element_located((By.CSS_SELECTOR,".Review-paginator.Review-paginator--withSteps")))
        soup_first = BeautifulSoup(driver.page_source,"html.parser")
        comments_first = soup_first.findAll("div",class_="Review-comments")
        write_comments(comments_first,"tw")
        ransleep()
        page_list = count_page()
        if page_list > 1:
            ele = driver.find_element(By.CLASS_NAME,"Review-filters")
            driver.execute_script("arguments[0].scrollIntoView()",ele)
            ransleep()
            page_click("tw")
    sleep(1)
    
    if korea_btn:
        filter_btn.click()
        korea_btn.click()
        WebDriverWait(driver,5).until(EC.presence_of_element_located((By.CSS_SELECTOR,".Review-paginator.Review-paginator--withSteps")))
        soup_first = BeautifulSoup(driver.page_source,"html.parser")
        write_comments(soup_first,"kr")
        ransleep()
        page_list = count_page()
        if page_list > 1:
            ele = driver.find_element(By.CLASS_NAME,"Review-filters")
            driver.execute_script("arguments[0].scrollIntoView()",ele)
            ransleep()
            page_click("kr")
    sleep(1)
    
    if japan_btn:
        filter_btn.click()
        japan_btn.click()
        WebDriverWait(driver,5).until(EC.presence_of_element_located((By.CSS_SELECTOR,".Review-paginator.Review-paginator--withSteps")))
        soup_first = BeautifulSoup(driver.page_source,"html.parser")
        write_comments(soup_first,"jp")
        ransleep()
        try:
            ele_page_count = driver.find_element(By.CLASS_NAME,"Review__SummaryContainer")
        except:
            ele_page_count = ""
        page_list = count_page()
        if page_list > 1:
            ele = driver.find_element(By.CLASS_NAME,"Review-filters")
            driver.execute_script("arguments[0].scrollIntoView()",ele)
            ransleep()
            page_click("jp")

driver.close()
print("finished")


                